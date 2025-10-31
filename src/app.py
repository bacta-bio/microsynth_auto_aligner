#!/usr/bin/env python3
"""
Flask web server for Microsynth Auto Aligner
"""
from flask import Flask, render_template, request, jsonify
import os
import tempfile
import zipfile
import shutil
import sys

# Add the parent directory to the Python path so we can import from src
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from src.microsynth_auto_aligner import run_alignment, set_log_function
from benchling.client import BenchlingClient
import pandas as pd
from io import BytesIO
import re
import openpyxl  # for Eurofins export
try:
    import docker  # docker SDK for helper logs
except Exception:
    docker = None

# Get the absolute path to the templates and static directories
template_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'templates'))
static_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'static'))

app = Flask(__name__, template_folder=template_dir, static_folder=static_dir)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max upload size
app.config['UPLOAD_FOLDER'] = '/tmp/uploads'  # Use persistent temp directory

# Store logs and alignment results in memory for this session
logs_buffer = []
alignment_results = []

# Initialize Benchling client (OAuth2)
benchling_client = BenchlingClient()

# Primer Registration constants (from legacy Dash app)
REGISTRY_ID = 'src_LW5X8lCL'
DROPDOWN_ID = 'sfs_kKhZZg7c'  # Direction dropdown
SCHEMA_ID = 'ts_lFefhdfz'

def web_log(message: str) -> None:
    """Log function that sends messages to the web interface."""
    logs_buffer.append(message)
    print(message)  # Also print to console
    # Keep only last 100 log messages
    if len(logs_buffer) > 100:
        logs_buffer.pop(0)

@app.route('/')
def index():
    """Main page"""
    return render_template('index.html')

@app.route('/health', methods=['GET'])
def health():
    """Simple health endpoint for load balancers."""
    return jsonify({"status": "ok"}), 200

@app.route('/healthz', methods=['GET'])
def healthz():
    """Kubernetes/Compose-friendly health endpoint alias."""
    return jsonify({"status": "ok"}), 200

@app.route('/api/logs')
def get_logs():
    """Get recent log messages"""
    return jsonify({'logs': logs_buffer})

@app.route('/api/results')
def get_results():
    """Get alignment results with Benchling links"""
    return jsonify({'results': alignment_results})

@app.route('/api/users', methods=['GET'])
def list_users():
    """List Benchling users for assignment (name, id)."""
    if benchling_client is None:
        return jsonify({"error": "Benchling client not initialized"}), 500
    try:
        resp = benchling_client.make_request('GET', '/users')
        data = resp.json()
        users = data.get('users', [])
        mapped = [{"label": u.get('name'), "value": u.get('id')} for u in users if u.get('name') and u.get('id')]
        return jsonify({"users": mapped})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/dropdown/options', methods=['GET'])
def dropdown_options():
    """Return dropdown options for Direction field."""
    if benchling_client is None:
        return jsonify({"error": "Benchling client not initialized"}), 500
    try:
        # use helper to fetch dropdown options
        from benchling.client import BenchlingClient as _BC  # type: ignore
        # our client has get_dropdown_options method
        options = benchling_client.get_dropdown_options(SCHEMA_ID, 'Direction') if hasattr(benchling_client, 'get_dropdown_options') else {}
        # If helper unavailable, fall back to direct dropdowns endpoint
        if not options:
            resp_schema = benchling_client.make_request('GET', f'/entity-schemas/{SCHEMA_ID}')
            dropdown_id = None
            for fd in resp_schema.json().get('fieldDefinitions', []):
                if fd.get('name') == 'Direction' and fd.get('type') == 'dropdown':
                    dropdown_id = fd.get('dropdownId')
                    break
            if dropdown_id:
                resp_dd = benchling_client.make_request('GET', f'/dropdowns/{dropdown_id}')
                options = {opt['name']: opt['id'] for opt in resp_dd.json().get('options', [])}
        # Map to simple list
        items = [{'label': k, 'value': v} for k, v in options.items()]
        return jsonify({'options': items})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/upload', methods=['POST'])
def upload_files():
    """Handle file upload and extract if needed"""
    if 'files' not in request.files:
        return jsonify({'error': 'No files uploaded'}), 400
    
    files = request.files.getlist('files')
    if not files or files[0].filename == '':
        return jsonify({'error': 'No files selected'}), 400
    
    # Create a temporary directory for this upload
    upload_dir = tempfile.mkdtemp(dir=app.config['UPLOAD_FOLDER'])
    
    try:
        for file in files:
            filename = file.filename
            
            # Check if it's a zip file
            if filename.lower().endswith('.zip'):
                # Save and extract zip
                zip_path = os.path.join(upload_dir, filename)
                file.save(zip_path)
                
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(upload_dir)
                
                os.remove(zip_path)
            else:
                # Save individual file
                file.save(os.path.join(upload_dir, filename))
        
        return jsonify({
            'success': True,
            'upload_dir': upload_dir,
            'message': f'Successfully uploaded {len(files)} file(s)'
        })
    except Exception as e:
        shutil.rmtree(upload_dir, ignore_errors=True)
        return jsonify({'error': f'Error processing upload: {str(e)}'}), 500

@app.route('/api/primer/preview', methods=['POST'])
def primer_preview():
    """Upload CSV and return parsed rows for preview."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    f = request.files['file']
    if f.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    try:
        df = pd.read_csv(f)
    except Exception:
        try:
            f.stream.seek(0)
            df = pd.read_excel(f)
        except Exception as e:
            return jsonify({'error': f'Unable to parse file: {e}'}), 400
    required = ['Name', 'Sequence', 'Description']
    for col in required:
        if col not in df.columns:
            return jsonify({'error': f"Missing required column: {col}"}), 400
    rows = df[required].fillna("").to_dict(orient='records')
    return jsonify({'rows': rows, 'count': len(rows)})

@app.route('/api/primer/register', methods=['POST'])
def primer_register():
    """Register primers to Benchling from provided rows."""
    if benchling_client is None:
        return jsonify({"error": "Benchling client not initialized"}), 500
    data = request.get_json(silent=True) or {}
    user_id = data.get('userId')
    registry_id = REGISTRY_ID
    schema_id = SCHEMA_ID
    rows = data.get('rows', [])
    if not user_id or not registry_id or not schema_id or not rows:
        return jsonify({'error': 'userId, registryId, schemaId and rows are required'}), 400
    results = []
    for row in rows:
        try:
            payload = {
                "name": str(row.get('Name', '')),
                "registryId": registry_id,
                "schemaId": schema_id,
                "namingStrategy": "NEW_IDS",
                "bases": str(row.get('Sequence', '')),
                "authorIds": [user_id],
                "fields": {
                    "Description": {"textValue": str(row.get('Description', ''))}
                }
            }
            resp = benchling_client.make_request('POST', '/dna-oligos', data=payload)
            rj = resp.json()
            # Post-create rename to NUMERICID_original
            new_name = rj.get('name') or payload['name']
            try:
                ent_id = rj.get('id')
                # fetch details to get entityRegistryId
                det = benchling_client.make_request('GET', f'/dna-oligos/{ent_id}').json()
                eri = det.get('entityRegistryId') or det.get('entity_registry_id') or ''
                digits = ''.join(re.findall(r'\d+', str(eri)))
                if digits:
                    patched_name = f"{digits}_{payload['name']}"
                    benchling_client.make_request('PATCH', f"/dna-oligos/{ent_id}", data={"name": patched_name})
                    new_name = patched_name
            except Exception:
                pass
            results.append({
                'Oligo Name': new_name,
                'Sequence': payload['bases'],
                'Personal Note': row.get('Description', '')
            })
        except Exception as e:
            results.append({
                'Oligo Name': row.get('Name', ''),
                'Sequence': row.get('Sequence', ''),
                'Personal Note': f'ERROR: {e}'
            })
    return jsonify({'results': results, 'count': len(results)})

@app.route('/api/primer/eurofins', methods=['POST'])
def primer_eurofins():
    """Generate Eurofins Excel from results using template like legacy Dash app."""
    if openpyxl is None:
        return jsonify({'error': 'openpyxl not installed in container'}), 500
    data = request.get_json(silent=True) or {}
    rows = data.get('rows', [])
    template_path = os.path.join(os.path.dirname(__file__), '..', 'data', 'eurofins_upload-template_customdnaoligos.xlsx')
    template_path = os.path.abspath(template_path)
    if not os.path.exists(template_path):
        return jsonify({'error': f'Template not found at {template_path}'}), 500
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb['Form']
        # Write starting at row 9: col2 Name, col3 Sequence, col4 Personal Note
        r = 9
        for row in rows:
            ws.cell(row=r, column=2, value=str(row.get('Oligo Name', ''))[:25])
            ws.cell(row=r, column=3, value=str(row.get('Sequence', '')))
            ws.cell(row=r, column=4, value=str(row.get('Personal Note', '')))
            r += 1
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        from flask import send_file
        return send_file(bio, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='bacta_eurofins_registered_primers.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/benchling-helper/logs', methods=['GET'])
def benchling_helper_logs():
    """Fetch recent logs from the 'benchling-helper' Docker container.
    Requires docker socket mounted and docker SDK available.
    """
    tail = int(request.args.get('tail', 200))
    if docker is None:
        return jsonify({'error': 'Docker SDK not available in app container'}), 500
    try:
        client = docker.from_env()
        container = client.containers.get('benchling-helper')
        log_bytes = container.logs(tail=tail)
        text = log_bytes.decode('utf-8', errors='replace')
        lines = text.strip().split('\n') if text else []
        return jsonify({'lines': lines[-tail:]})
    except Exception as e:
        return jsonify({'error': f'Unable to read logs: {e}'}), 500

@app.route('/api/run', methods=['POST'])
def run_alignment_api():
    """Run the alignment process"""
    global logs_buffer, alignment_results
    logs_buffer = []  # Clear logs
    alignment_results = []  # Clear previous results
    
    data = request.json
    upload_dir = data.get('upload_dir', '')
    
    if not upload_dir:
        return jsonify({'error': 'No upload directory provided'}), 400
    
    if not os.path.exists(upload_dir):
        return jsonify({'error': 'Upload directory does not exist'}), 400
    
    try:
        # Set up custom log function
        set_log_function(web_log)
        
        # Run the alignment
        success, results = run_alignment(upload_dir)
        alignment_results = results
        
        return jsonify({
            'success': success,
            'message': 'Alignment completed successfully' if success else 'Alignment failed',
            'results_count': len(results)
        })
    finally:
        # Clean up temporary files
        if upload_dir and os.path.exists(upload_dir):
            shutil.rmtree(upload_dir, ignore_errors=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080, debug=False)

